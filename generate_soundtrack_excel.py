#!/usr/bin/env python3
"""
Generate an Excel file with soundtrack information from sample directory.

The Excel file contains:
- movie_id: name of directory in sample directory
- movie name: extracted from movie_html/<id>.ttl file
- track name: from soundtrack_links.json
- track url: YouTube URL from best_match
- correct (y/N): empty column for manual validation
- comments: extracted from match_score reasoning/concerns or error messages
"""
import json
import sys
from pathlib import Path
from typing import List, Dict, Any, Optional

from rdflib import Graph, URIRef, Namespace
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def extract_movie_name(movie_dir: Path, movie_id: str) -> str:
    """Extract movie name from movie_html/<id>.ttl file."""
    movie_ttl_path = movie_dir / 'movie_html' / f'{movie_id}.ttl'
    
    if not movie_ttl_path.exists():
        print(f"Warning: Movie TTL not found: {movie_ttl_path}", file=sys.stderr)
        return ''
    
    try:
        schema = Namespace("http://schema.org/")
        g = Graph()
        g.parse(movie_ttl_path.as_posix(), format="turtle")
        
        # Try both URI variants (with and without trailing slash)
        movie_uri_no_slash = URIRef(f"https://www.imdb.com/title/{movie_id}")
        movie_uri_slash = URIRef(f"https://www.imdb.com/title/{movie_id}/")
        
        # Find movie name
        for movie_uri in [movie_uri_no_slash, movie_uri_slash]:
            names = list(g.objects(movie_uri, schema.name))
            if names:
                return str(names[0])
        
        return ''
    except Exception as e:
        print(f"Error: Failed to parse movie TTL {movie_ttl_path}: {e}", file=sys.stderr)
        return ''


def load_soundtrack_json(json_path: Path) -> List[Dict[str, Any]]:
    """Load soundtrack_links.json file."""
    if not json_path.exists():
        print(f"Warning: {json_path} does not exist", file=sys.stderr)
        return []
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        print(f"Error: Failed to parse JSON in {json_path}: {e}", file=sys.stderr)
        return []
    except Exception as e:
        print(f"Error: Failed to read {json_path}: {e}", file=sys.stderr)
        return []


def extract_track_info(track_data: Dict[str, Any]) -> Dict[str, str]:
    """Extract track information from a single track entry."""
    soundtrack = track_data.get('soundtrack', {})
    track_name = soundtrack.get('title', '')
    
    # Extract URL
    track_url = ''
    if 'best_match' in track_data and track_data['best_match']:
        track_url = track_data['best_match'].get('url', '')
    
    # Extract comments
    comments_parts = []
    
    # Add error message if present
    if 'error' in track_data and track_data['error']:
        comments_parts.append(f"Error: {track_data['error']}")
    
    # Add match score reasoning if present
    if 'match_score' in track_data and track_data['match_score']:
        match_score = track_data['match_score']
        if 'reasoning' in match_score and match_score['reasoning']:
            comments_parts.append(f"Reasoning: {match_score['reasoning']}")
        
        # Add concerns if present
        if 'concerns' in match_score and match_score['concerns']:
            concerns = match_score['concerns']
            if isinstance(concerns, list):
                concerns_str = '; '.join(concerns)
            else:
                concerns_str = str(concerns)
            if concerns_str:
                comments_parts.append(f"Concerns: {concerns_str}")
    
    comments = ' | '.join(comments_parts) if comments_parts else ''
    
    return {
        'track_name': track_name,
        'track_url': track_url,
        'comments': comments
    }


def generate_excel(sample_dir: Path, output_file: Path):
    """Generate Excel file from sample directory."""
    if not sample_dir.exists():
        print(f"Error: Sample directory does not exist: {sample_dir}", file=sys.stderr)
        sys.exit(1)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Soundtrack Links"
    
    # Set headers
    headers = ['movie_id', 'movie name', 'track name', 'track url', 'correct (y/N)', 'comments']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Process each movie directory
    movie_dirs = sorted([d for d in sample_dir.iterdir() if d.is_dir()])
    
    if not movie_dirs:
        print(f"Warning: No directories found in {sample_dir}", file=sys.stderr)
    
    total_rows = 0
    
    for movie_dir in movie_dirs:
        movie_id = movie_dir.name
        json_path = movie_dir / 'movie_soundtrack' / 'soundtrack_links.json'
        
        if not json_path.exists():
            print(f"Warning: {json_path} does not exist, skipping {movie_id}", file=sys.stderr)
            continue
        
        # Extract movie name from TTL file
        movie_name = extract_movie_name(movie_dir, movie_id)
        
        # Load soundtrack data
        tracks = load_soundtrack_json(json_path)
        
        if not tracks:
            print(f"Warning: No tracks found in {json_path}, skipping {movie_id}", file=sys.stderr)
            continue
        
        # Add a row for each track
        for track_data in tracks:
            track_info = extract_track_info(track_data)
            
            row = [
                movie_id,
                movie_name,
                track_info['track_name'],
                track_info['track_url'],
                '',  # correct (y/N) - empty for manual filling
                track_info['comments']
            ]
            ws.append(row)
            total_rows += 1
    
    # Auto-adjust column widths
    column_widths = {
        'A': 15,  # movie_id
        'B': 30,  # movie name
        'C': 50,  # track name
        'D': 50,  # track url
        'E': 15,  # correct (y/N)
        'F': 100  # comments
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Wrap text in comments column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Save workbook
    wb.save(output_file)
    print(f"✓ Generated Excel file: {output_file}")
    print(f"  Total rows: {total_rows}")
    print(f"  Movies processed: {len(movie_dirs)}")


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Generate Excel file with soundtrack information from sample directory"
    )
    parser.add_argument(
        '--sample-dir',
        type=str,
        default='data/sample',
        help='Path to sample directory containing movie subdirectories'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='soundtrack_links.xlsx',
        help='Output Excel filename'
    )
    
    args = parser.parse_args()
    
    sample_dir = Path(args.sample_dir).resolve()
    output_file = Path(args.output).resolve()
    
    generate_excel(sample_dir, output_file)


if __name__ == '__main__':
    main()

