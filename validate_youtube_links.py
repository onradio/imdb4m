#!/usr/bin/env python3
"""
Interactive YouTube link validator with GUI.

Opens YouTube links from Excel file one by one in Chrome browser,
displays a GUI window for validation, and updates the Excel file
with Y/N responses.
"""
import sys
import time
import threading
from pathlib import Path
from typing import Optional, Dict, Any
from dataclasses import dataclass

# Import tkinter only when needed (checked in run method)
try:
    import tkinter as tk
    from tkinter import ttk, messagebox
    TKINTER_AVAILABLE = True
except ImportError:
    TKINTER_AVAILABLE = False

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook


@dataclass
class TrackRow:
    """Represents a row from the Excel file."""
    row_index: int  # Excel row number (1-indexed, including header)
    movie_id: str
    movie_name: str
    track_name: str
    track_url: str
    correct: str  # Current value in correct (y/N) column
    comments: str


class YouTubeValidator:
    """Main validator class that manages browser, GUI, and Excel updates."""
    
    def __init__(self, excel_path: Path, interval_seconds: int = 20):
        self.excel_path = excel_path
        self.interval_seconds = interval_seconds
        self.driver: Optional[webdriver.Chrome] = None
        self.workbook = None
        self.worksheet = None
        self.tracks: list[TrackRow] = []
        self.current_index = 0
        self.auto_advance_timer: Optional[threading.Timer] = None
        
        # GUI components
        self.root: Optional[tk.Tk] = None
        self.movie_label: Optional[tk.Label] = None
        self.track_label: Optional[tk.Label] = None
        self.progress_label: Optional[tk.Label] = None
        self.yes_button: Optional[tk.Button] = None
        self.no_button: Optional[tk.Button] = None
        self.skip_button: Optional[tk.Button] = None
        
    def load_excel(self):
        """Load Excel file and extract track data."""
        print(f"Loading Excel file: {self.excel_path}")
        self.workbook = load_workbook(self.excel_path)
        self.worksheet = self.workbook.active
        
        # Read all rows (skip header row)
        self.tracks = []
        for row_idx in range(2, self.worksheet.max_row + 1):
            movie_id = self.worksheet.cell(row=row_idx, column=1).value or ''
            movie_name = self.worksheet.cell(row=row_idx, column=2).value or ''
            track_name = self.worksheet.cell(row=row_idx, column=3).value or ''
            track_url = self.worksheet.cell(row=row_idx, column=4).value or ''
            correct = self.worksheet.cell(row=row_idx, column=5).value or ''
            comments = self.worksheet.cell(row=row_idx, column=6).value or ''
            
            # Only include rows with URLs
            if track_url:
                self.tracks.append(TrackRow(
                    row_index=row_idx,
                    movie_id=str(movie_id),
                    movie_name=str(movie_name),
                    track_name=str(track_name),
                    track_url=str(track_url),
                    correct=str(correct),
                    comments=str(comments)
                ))
        
        print(f"Loaded {len(self.tracks)} tracks with URLs")
        
        # Filter out already validated tracks (only process tracks without Y or N)
        validated_count = len(self.tracks)
        self.tracks = [
            t for t in self.tracks 
            if not t.correct.strip() or t.correct.strip().upper() not in ['Y', 'N']
        ]
        print(f"After filtering validated tracks: {len(self.tracks)} remaining (skipped {validated_count - len(self.tracks)} already validated)")
    
    def setup_browser(self):
        """Set up Chrome browser with Selenium."""
        print("Setting up Chrome browser...")
        chrome_options = Options()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        # Selenium Manager will locate the correct ChromeDriver automatically
        # (built into Selenium 4.15.2+, no need for webdriver-manager)
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.maximize_window()
        print("✓ Browser ready")
    
    def create_gui(self):
        """Create the GUI window."""
        self.root = tk.Tk()
        self.root.title("YouTube Link Validator")
        self.root.geometry("500x350")
        self.root.resizable(False, False)
        
        # Center window on screen
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
        # Make window always on top
        self.root.attributes('-topmost', True)
        
        # Add keyboard shortcuts
        self.root.bind('<y>', lambda e: self.handle_response('Y'))
        self.root.bind('<Y>', lambda e: self.handle_response('Y'))
        self.root.bind('<n>', lambda e: self.handle_response('N'))
        self.root.bind('<N>', lambda e: self.handle_response('N'))
        self.root.bind('<s>', lambda e: self.skip_current())
        self.root.bind('<S>', lambda e: self.skip_current())
        self.root.bind('<space>', lambda e: self.skip_current())
        
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Progress label
        self.progress_label = ttk.Label(
            main_frame,
            text="",
            font=("Arial", 10, "bold")
        )
        self.progress_label.pack(pady=(0, 20))
        
        # Movie name label
        movie_frame = ttk.Frame(main_frame)
        movie_frame.pack(fill=tk.X, pady=10)
        ttk.Label(movie_frame, text="Movie:", font=("Arial", 12, "bold")).pack(side=tk.LEFT, padx=5)
        self.movie_label = ttk.Label(
            movie_frame,
            text="",
            font=("Arial", 12),
            foreground="blue"
        )
        self.movie_label.pack(side=tk.LEFT, padx=5)
        
        # Track name label
        track_frame = ttk.Frame(main_frame)
        track_frame.pack(fill=tk.X, pady=10)
        ttk.Label(track_frame, text="Track:", font=("Arial", 12, "bold")).pack(side=tk.LEFT, padx=5)
        self.track_label = ttk.Label(
            track_frame,
            text="",
            font=("Arial", 11),
            foreground="darkgreen",
            wraplength=400
        )
        self.track_label.pack(side=tk.LEFT, padx=5)
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=30)
        
        self.yes_button = tk.Button(
            button_frame,
            text="✓ YES",
            font=("Arial", 14, "bold"),
            bg="#4CAF50",
            fg="white",
            width=10,
            height=2,
            command=lambda: self.handle_response('Y')
        )
        self.yes_button.pack(side=tk.LEFT, padx=10)
        
        self.no_button = tk.Button(
            button_frame,
            text="✗ NO",
            font=("Arial", 14, "bold"),
            bg="#f44336",
            fg="white",
            width=10,
            height=2,
            command=lambda: self.handle_response('N')
        )
        self.no_button.pack(side=tk.LEFT, padx=10)
        
        self.skip_button = tk.Button(
            button_frame,
            text="Skip",
            font=("Arial", 12),
            bg="#ff9800",
            fg="white",
            width=8,
            height=2,
            command=self.skip_current
        )
        self.skip_button.pack(side=tk.LEFT, padx=10)
        
        # Status label
        self.status_label = ttk.Label(
            main_frame,
            text="",
            font=("Arial", 9),
            foreground="gray"
        )
        self.status_label.pack(pady=10)
        
        # Keyboard shortcuts hint
        hint_label = ttk.Label(
            main_frame,
            text="Keyboard shortcuts: Y (Yes), N (No), S/Space (Skip)",
            font=("Arial", 8),
            foreground="darkgray"
        )
        hint_label.pack(pady=5)
        
        # Handle window close
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def update_gui(self, track: TrackRow):
        """Update GUI with current track information."""
        if not self.root:
            return
        
        # Update progress
        self.progress_label.config(
            text=f"Track {self.current_index + 1} of {len(self.tracks)}"
        )
        
        # Update movie and track names
        self.movie_label.config(text=track.movie_name or track.movie_id)
        self.track_label.config(text=track.track_name or "(No track name)")
        
        # Update status
        current_status = f"Current: {track.correct}" if track.correct.strip() else "Not validated"
        self.status_label.config(text=current_status)
    
    def open_url(self, url: str):
        """Open URL in browser."""
        if not self.driver:
            return
        
        try:
            print(f"Opening URL: {url}")
            self.driver.get(url)
            time.sleep(2)  # Wait for page to load
        except Exception as e:
            print(f"Error opening URL: {e}")
            messagebox.showerror("Error", f"Failed to open URL:\n{e}")
    
    def schedule_auto_advance(self):
        """Schedule automatic advance to next track after interval."""
        if self.auto_advance_timer:
            self.auto_advance_timer.cancel()
        
        self.auto_advance_timer = threading.Timer(
            self.interval_seconds,
            self.auto_advance
        )
        self.auto_advance_timer.start()
    
    def auto_advance(self):
        """Automatically advance to next track."""
        if self.current_index < len(self.tracks) - 1:
            self.current_index += 1
            self.load_current_track()
        else:
            self.status_label.config(text="All tracks processed!")
    
    def load_current_track(self):
        """Load and display current track."""
        if self.current_index >= len(self.tracks):
            messagebox.showinfo("Complete", "All tracks have been processed!")
            return
        
        track = self.tracks[self.current_index]
        self.update_gui(track)
        
        # Open URL in browser
        if track.track_url:
            self.open_url(track.track_url)
        
        # Schedule auto-advance
        self.schedule_auto_advance()
    
    def update_excel(self, track: TrackRow, value: str):
        """Update Excel file with validation result."""
        try:
            # Update the correct (y/N) column (column 5)
            self.worksheet.cell(row=track.row_index, column=5).value = value
            self.workbook.save(self.excel_path)
            print(f"✓ Updated row {track.row_index} with '{value}'")
            track.correct = value
        except Exception as e:
            print(f"Error updating Excel: {e}")
            messagebox.showerror("Error", f"Failed to update Excel file:\n{e}")
    
    def handle_response(self, value: str):
        """Handle Yes/No button click."""
        if self.auto_advance_timer:
            self.auto_advance_timer.cancel()
        
        track = self.tracks[self.current_index]
        self.update_excel(track, value)
        
        # Update status
        self.status_label.config(text=f"Saved: {value}")
        
        # Advance to next track
        if self.current_index < len(self.tracks) - 1:
            self.current_index += 1
            self.load_current_track()
        else:
            self.status_label.config(text="All tracks processed!")
            messagebox.showinfo("Complete", "All tracks have been processed!")
    
    def skip_current(self):
        """Skip current track without updating."""
        if self.auto_advance_timer:
            self.auto_advance_timer.cancel()
        
        # Advance to next track
        if self.current_index < len(self.tracks) - 1:
            self.current_index += 1
            self.load_current_track()
        else:
            self.status_label.config(text="All tracks processed!")
            messagebox.showinfo("Complete", "All tracks have been processed!")
    
    def on_closing(self):
        """Handle window close event."""
        if self.auto_advance_timer:
            self.auto_advance_timer.cancel()
        
        # Save workbook before closing
        try:
            if self.workbook:
                self.workbook.save(self.excel_path)
                print("✓ Excel file saved")
        except Exception as e:
            print(f"Error saving Excel: {e}")
        
        # Close browser
        if self.driver:
            self.driver.quit()
            print("Browser closed")
        
        # Close GUI
        if self.root:
            self.root.destroy()
    
    def run(self):
        """Run the validator."""
        # Check for tkinter
        if not TKINTER_AVAILABLE:
            print("Error: tkinter is not installed.")
            print("On Ubuntu/Debian, install it with: sudo apt-get install python3-tk")
            print("On Fedora/RHEL, install it with: sudo dnf install python3-tkinter")
            sys.exit(1)
        
        try:
            # Load Excel file
            self.load_excel()
            
            if not self.tracks:
                print("No tracks with URLs found in Excel file!")
                return
            
            # Setup browser
            self.setup_browser()
            
            # Create GUI
            self.create_gui()
            
            # Load first track
            self.load_current_track()
            
            # Start GUI event loop
            self.root.mainloop()
            
        except KeyboardInterrupt:
            print("\nInterrupted by user")
            self.on_closing()
        except Exception as e:
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()
            self.on_closing()


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Interactive YouTube link validator with GUI"
    )
    parser.add_argument(
        '--excel-file',
        type=str,
        default='soundtrack_links.xlsx',
        help='Path to Excel file with soundtrack links'
    )
    parser.add_argument(
        '--interval',
        type=int,
        default=20,
        help='Interval in seconds between automatic track advances (default: 20)'
    )
    
    args = parser.parse_args()
    
    excel_path = Path(args.excel_file).resolve()
    if not excel_path.exists():
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)
    
    validator = YouTubeValidator(excel_path, interval_seconds=args.interval)
    validator.run()


if __name__ == '__main__':
    main()

