"""Emit a human-readable Excel audit of the cleanup manifest."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .config import REPORT_PATH
from .reconcile import Action, Manifest

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_ACTION_FILL = {
    "keep":    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "rewrite": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "delete":  PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}


def write_report(manifest: Manifest, path: Path = REPORT_PATH) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ---- Summary sheet ------------------------------------------------------
    ws = wb.active
    ws.title = "summary"
    ws.append(["media_type", "keep", "rewrite", "delete", "total"])
    for col in ws[1]:
        col.font = _HEADER_FONT
        col.fill = _HEADER_FILL
    for mt in ("images", "videos", "audio"):
        s = manifest.stats.get(mt, {})
        k, r, d = s.get("keep", 0), s.get("rewrite", 0), s.get("delete", 0)
        ws.append([mt, k, r, d, k + r + d])

    # ---- Detail sheets ------------------------------------------------------
    cols = ["media_type", "action", "entity_id", "old_uri", "new_uri",
            "old_url", "new_url", "audio_title", "audio_file_on_disk",
            "audio_bnode", "reason"]
    for action_name in ("rewrite", "delete"):
        ws = wb.create_sheet(action_name)
        ws.append(cols)
        for col in ws[1]:
            col.font = _HEADER_FONT
            col.fill = _HEADER_FILL
        for d in manifest.decisions:
            if d.action.value != action_name:
                continue
            row = [getattr(d, c, "") or "" for c in cols]
            row[1] = d.action.value
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.fill = _ACTION_FILL[action_name]
        for column_cells in ws.columns:
            width = min(max(len(str(c.value or "")) for c in column_cells) + 2, 80)
            ws.column_dimensions[column_cells[0].column_letter].width = width

    wb.save(str(path))
    logger.info("Wrote report → %s", path)
    return path
